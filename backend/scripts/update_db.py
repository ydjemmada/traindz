import os
import sys
import django
from openpyxl import load_workbook
from datetime import time as dt_time

# Setup Django environment
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "sntf_project.settings")
django.setup()

from api.models import Station, Route, Train, Stop, Line, Connection

# Station name normalization mapping
STATION_NAME_MAPPING = {
    'El-Harrach': 'El Harrach',
    'Oued-Smar': 'Oued Smar',
    'Bab-Ezzouar': 'Bab Ezzouar',
    'Dar-El-Beida': 'Dar El Beida',
    'Dar El-Beida': 'Dar El Beida',
    'Dar El Beïda': 'Dar El Beida',
    'Rouiba-Ind': 'Rouiba Ind',
    'Rouiba-SNVI': 'Rouiba SNVI',
    'Reghaia-Ind': 'Reghaia Ind',
    'Gué-de-Constantine': 'Gué de Constantine',
    'Ain-Naadja': 'Ain Naadja',
    'Baba-Ali': 'Baba Ali',
    'Beni-Mered': 'Beni Mered',
    'Tessala-El-Merdja': 'Tessala El Merdja',
    'Sidi-Abdelah': 'Sidi Abdelah',
    'Aeroport-Houari-Boumediene': 'Aeroport Houari Boumediene',
    'Thénia': 'Thenia',
    'Boumerdès': 'Boumerdes',
    'Réghaïa': 'Reghaia',
    'H.Dey': 'Hussein Dey',
    'Sidi Abde allah': 'Sidi Abdelah',
    'Sidi Abde allah-U': 'Sidi Abdelah-U',
}

def normalize_station_name(name):
    """Normalize station name to canonical form"""
    if not name:
        return None
    name = ' '.join(name.split())
    return STATION_NAME_MAPPING.get(name, name)

def time_to_string(time_value):
    """Convert time value to HH:MM string"""
    if isinstance(time_value, str):
        if ':' in time_value:
            return time_value.strip()
        return None
    elif isinstance(time_value, dt_time):
        return f"{time_value.hour:02d}:{time_value.minute:02d}"
    elif hasattr(time_value, 'hour'):
        return f"{time_value.hour:02d}:{time_value.minute:02d}"
    return None

def parse_operating_days(days_str):
    """Parse operating days string to Enum value"""
    if not days_str:
        return 'daily'
    
    days_str = str(days_str).strip()
    
    if '[*]' in days_str:
        return 'daily'
    elif '[1]' in days_str:
        return 'no_friday'
    elif '[2]' in days_str:
        return 'friday_only'
    
    return 'daily' # Default

def update_database(file_path):
    """Update database from SNTF Excel file"""
    
    print(f"📂 Reading Excel file: {file_path}")
    
    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"❌ Error loading Excel file: {e}")
        return
    
    # Get all existing stations
    station_objs = {s.name_fr: s for s in Station.objects.all()}
    route_map = {r.name: r for r in Route.objects.all()}
    
    # Sheet name to route mapping
    sheet_to_route = {
        'Alger-Thenia': 'Alger - Thenia',
        'Thenia-Alger': 'Thenia - Alger',
        'Alger-OuedAissi': 'Alger - Thenia',
        'Alger-El Affroun': 'Alger - El Affroun',
        'El Affroun-Alger': 'El Affroun - Alger',
        'Thenia-ElAfroun': 'Alger - El Affroun',
        'Alger-Zeralda': 'Alger - Zéralda',
        'Zeralda-Alger': 'Zéralda - Alger',
        'Thenia-Zeralda': 'Alger - Zéralda',
        'Alger-Blida': 'Alger - El Affroun',
        'Alger-Oran': 'Alger - Oran',
        'Oran-Alger': 'Oran - Alger',
    }
    
    imported_count = 0
    skipped_count = 0
    
    # Clear existing trains to avoid duplicates/stale data
    print("🧹 Clearing existing schedule data...")
    Train.objects.all().delete()
    
    for sheet_name in wb.sheetnames:
        print(f"\n📄 Processing sheet: {sheet_name}")
        ws = wb[sheet_name]
        
        # Determine route name
        route_name = None
        for sn, rn in sheet_to_route.items():
            if sn.lower().replace('-', '').replace(' ', '') == sheet_name.lower().replace('-', '').replace(' ', ''):
                route_name = rn
                break
        
        if not route_name:
            print(f"   ⚠️  Could not map sheet '{sheet_name}' to a route")
            continue
            
        # Ensure Route Exists
        if route_name not in route_map:
            print(f"   🛠️  Creating missing route: {route_name}")
            try:
                origin_name, dest_name = route_name.split(' - ')
                
                # Ensure Origin/Dest Stations exist
                for s_name in [origin_name, dest_name]:
                    if s_name not in station_objs:
                        print(f"      Creating station: {s_name}")
                        s = Station.objects.create(name_fr=s_name, name_ar=s_name)
                        station_objs[s_name] = s
                
                # Ensure Line exists
                line_name = route_name # Simplified: Route name = Line name
                line, _ = Line.objects.get_or_create(name=line_name)
                
                # Create Route
                route = Route.objects.create(
                    name=route_name,
                    line=line,
                    origin=station_objs[origin_name],
                    destination=station_objs[dest_name]
                )
                route_map[route_name] = route
            except Exception as e:
                print(f"   ❌ Error creating route {route_name}: {e}")
                continue

        route = route_map[route_name]
        print(f"   📍 Using route: {route.name}")
        
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 7:
            continue
        
        train_row = rows[4]  # Row 5
        days_row = rows[5]   # Row 6
        
        trains = []
        for col_idx in range(2, len(train_row)):
            train_num = train_row[col_idx]
            if train_num and train_num != 'Trains':
                days = days_row[col_idx] if col_idx < len(days_row) else '[*]'
                trains.append({
                    'number': str(train_num).strip(),
                    'col_idx': col_idx,
                    'days': str(days) if days else '[*]'
                })
        
        print(f"   Found {len(trains)} trains")
        
        for train_info in trains:
            train_number = train_info['number']
            col_idx = train_info['col_idx']
            days_str = train_info['days']
            
            operating_days_enum = parse_operating_days(days_str)
            
            train = Train.objects.create(
                number=train_number,
                route=route,
                days_operational=days_str,
                operating_days=operating_days_enum,
                active_status=True
            )
            
            stops_created = 0
            skipped_stations = []
            last_time_obj = None
            current_train = train
            split_count = 0
            last_station_name = None
            
            for row_idx in range(6, len(rows)):
                row = rows[row_idx]
                if len(row) <= 1: continue
                
                station_name = row[1]
                if not station_name or station_name == 'Gares\\Day': continue
                
                station_name = normalize_station_name(str(station_name).strip())
                time_value = row[col_idx] if col_idx < len(row) else None
                if not time_value: continue
                
                time_str = time_to_string(time_value)
                if not time_str or time_str == '-': continue
                
                if station_name not in station_objs:
                    # Create station on the fly
                    # print(f"      Creating new station: {station_name}")
                    s = Station.objects.create(name_fr=station_name, name_ar=station_name)
                    station_objs[station_name] = s
                
                # Split Logic
                is_new_trip = False
                try:
                    h, m = map(int, time_str.split(':'))
                    current_time_obj = dt_time(h, m)
                    
                    if last_time_obj:
                        if last_station_name and station_name == last_station_name:
                            is_new_trip = True
                        else:
                            curr_minutes = h * 60 + m
                            last_minutes = last_time_obj.hour * 60 + last_time_obj.minute
                            if curr_minutes < last_minutes:
                                is_new_trip = True
                    
                    last_time_obj = current_time_obj
                    last_station_name = station_name
                except:
                    pass

                if is_new_trip:
                    split_count += 1
                    new_number = f"{train_number}_{split_count + 1}"
                    current_train = Train.objects.create(
                        number=new_number,
                        route=route,
                        days_operational=days_str,
                        operating_days=operating_days_enum,
                        active_status=True
                    )
                    stops_created = 0

                Stop.objects.create(
                    train=current_train,
                    station=station_objs[station_name],
                    departure_time=time_str,
                    sequence=stops_created + 1
                )
                stops_created += 1
            
            if stops_created > 0:
                imported_count += 1
            else:
                skipped_count += 1
                current_train.delete() # Cleanup empty train

    print(f"\n✅ Update complete!")
    print(f"   Imported: {imported_count} trains")
    print(f"   Skipped: {skipped_count} trains")

if __name__ == "__main__":
    excel_path = "/home/mathxro/AntiGrav/SNTF_real.xlsx"
    print("=" * 60)
    print("  SNTF Automated Database Update")
    print("=" * 60)
    update_database(excel_path)
    print("\n" + "=" * 60)
