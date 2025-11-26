import os
import sys
import django
from openpyxl import load_workbook
from datetime import time as dt_time

# Setup Django environment
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "sntf_project.settings")
django.setup()

from api.models import Station, Route, Train, Stop, Line

# Station name normalization mapping
STATION_NAME_MAPPING = {
    # Variations with hyphens
    'El-Harrach': 'El Harrach',
    'Oued-Smar': 'Oued Smar',
    'Bab-Ezzouar': 'Bab Ezzouar',
    'Dar-El-Beida': 'Dar El Beida',
    'Dar El-Beida': 'Dar El Beida',
    'Dar El Beïda': 'Dar El Beida',
    'Rouiba-Ind': 'Rouiba Ind',
    'Rouiba-SNVI': 'Rouiba SNVI',
    'Reghaia-Ind': 'Reghaia Ind',
    'Gué-de-Constantine': 'Gué de Constantine',
    'Ain-Naadja': 'Ain Naadja',
    'Baba-Ali': 'Baba Ali',
    'Beni-Mered': 'Beni Mered',
    'Tessala-El-Merdja': 'Tessala El Merdja',
    'Sidi-Abdelah': 'Sidi Abdelah',
    'Aeroport-Houari-Boumediene': 'Aeroport Houari Boumediene',
    # Special characters (accents)
    'Thénia': 'Thenia',
    'Boumerdès': 'Boumerdes',
    'Réghaïa': 'Reghaia',
    # Abbreviations (these now exist as separate stations)
    # 'B.Mered' is kept separate
    # 'Gué de Cne' is kept separate
    'H.Dey': 'Hussein Dey',
    'Sidi Abde allah': 'Sidi Abdelah',
    'Sidi Abde allah-U': 'Sidi Abdelah-U',
}


def normalize_station_name(name):
    """Normalize station name to canonical form"""
    if not name:
        return None
    name = ' '.join(name.split())
    return STATION_NAME_MAPPING.get(name, name)

def time_to_string(time_value):
    """Convert time value to HH:MM string"""
    if isinstance(time_value, str):
        # It's already a string, might be a formula result or time string
        if ':' in time_value:
            return time_value.strip()
        return None
    elif isinstance(time_value, dt_time):
        # It's a datetime.time object
        return f"{time_value.hour:02d}:{time_value.minute:02d}"
    elif hasattr(time_value, 'hour'):
        # It's some datetime object
        return f"{time_value.hour:02d}:{time_value.minute:02d}"
    return None

def import_excel_trains(file_path):
    """Import trains from SNTF Excel file"""
    
    print(f"📂 Reading Excel file: {file_path}")
    
    try:
        # data_only=True to get calculated values instead of formulas
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"❌ Error loading Excel file: {e}")
        return
    
    # Get all existing stations
    station_objs = {s.name_fr: s for s in Station.objects.all()}
    route_map = {r.name: r for r in Route.objects.all()}
    
    # Sheet name to route mapping
    sheet_to_route = {
        'Alger-Thenia': 'Alger - Thenia',
        'Thenia-Alger': 'Thenia - Alger',
        'Alger-OuedAissi': 'Alger - Thenia',  # Extension of AT line
        'Alger-El Affroun': 'Alger - El Affroun',
        'El Affroun-Alger': 'El Affroun - Alger',
        'Thenia-ElAfroun': 'Alger - El Affroun',  # Connection
        'Alger-Zeralda': 'Alger - Zéralda',
        'Zeralda-Alger': 'Zéralda - Alger',
        'Thenia-Zeralda': 'Alger - Zéralda',  # Thenia to Zeralda via Alger
        'Alger-Blida': 'Alger - El Affroun',  # Subset of AE line
        'Alger-Oran': 'Alger - Oran',  # Intercity line
        'Oran-Alger': 'Oran - Alger',  # Intercity line
    }
    
    imported_count = 0
    skipped_count = 0
    
    for sheet_name in wb.sheetnames:
        print(f"\n📄 Processing sheet: {sheet_name}")
        ws = wb[sheet_name]
        
        # Determine route
        route = None
        for sn, rn in sheet_to_route.items():
            if sn.lower().replace('-', '').replace(' ', '') == sheet_name.lower().replace('-', '').replace(' ', ''):
                route = route_map.get(rn)
                break
        
        if not route:
            print(f"   ⚠️  Could not map sheet '{sheet_name}' to a route")
            continue
        
        print(f"   📍 Using route: {route.name}")
        
        # Get all rows
        rows = list(ws.iter_rows(values_only=True))
        
        if len(rows) < 7:
            print(f"   ⚠️  Sheet has insufficient data")
            continue
        
        # Row 5 (index 4) has train numbers
        # Row 6 (index 5) has operating days
        # Rows 7+ (index 6+) have stations and times
        
        train_row = rows[4]  # Row 5
        days_row = rows[5]   # Row 6
        
        # Extract train numbers starting from column C (index 2)
        trains = []
        for col_idx in range(2, len(train_row)):
            train_num = train_row[col_idx]
            if train_num and train_num != 'Trains':
                days = days_row[col_idx] if col_idx < len(days_row) else '[*]'
                trains.append({
                    'number': str(train_num).strip(),
                    'col_idx': col_idx,
                    'days': str(days) if days else '[*]'
                })
        
        print(f"   Found {len(trains)} trains: {[t['number'] for t in trains]}")
        
        # Process each train
        for train_info in trains:
            train_number = train_info['number']
            col_idx = train_info['col_idx']
            days = train_info['days']
            
            # Check if train exists
            existing_train = Train.objects.filter(number=train_number, route=route).first()
            if existing_train:
                train = existing_train
                train.stops.all().delete()
                print(f"   ℹ️  Updating train {train_number}")
            else:
                train = Train.objects.create(
                    number=train_number,
                    route=route,
                    days_operational=days
                )
                print(f"   ✅ Created train {train_number} ({days})")
            
            # Extract stops from rows 7+ (index 6+)
            stops_created = 0
            skipped_stations = []
            
            # Track previous time to detect new trips (e.g. return trip stacked in same column)
            last_time_obj = None
            current_train = train
            split_count = 0
            
            for row_idx in range(6, len(rows)):
                row = rows[row_idx]
                
                # Column B (index 1) has station names
                if len(row) <= 1:
                    continue
                
                station_name = row[1]
                if not station_name or station_name == 'Gares\\Day':
                    continue
                
                # Normalize station name
                station_name = normalize_station_name(str(station_name).strip())
                
                # Get time from the train's column
                time_value = row[col_idx] if col_idx < len(row) else None
                
                if not time_value:
                    continue
                
                time_str = time_to_string(time_value)
                
                if not time_str or time_str == '-':
                    continue
                
                # Find station in database
                if station_name not in station_objs:
                    if station_name not in skipped_stations:
                        skipped_stations.append(station_name)
                    continue
                
                # Check for time reset (indicating a new trip)
                is_new_trip = False
                
                # Parse current time
                try:
                    h, m = map(int, time_str.split(':'))
                    current_time_obj = dt_time(h, m)
                    
                    if last_time_obj:
                        # 1. Check for Turnaround (Same station as previous stop)
                        # We need to track the last station name for this
                        if 'last_station_name' in locals() and station_name == last_station_name:
                            is_new_trip = True
                            print(f"      🔄 Detected turnaround at {station_name}")
                        
                        # 2. Check for Time Drop (Current time < Last time)
                        # Unless it's a small adjustment or midnight crossing (unlikely for SNTF day trains)
                        else:
                            curr_minutes = h * 60 + m
                            last_minutes = last_time_obj.hour * 60 + last_time_obj.minute
                            
                            if curr_minutes < last_minutes:
                                is_new_trip = True
                                print(f"      🔄 Detected time drop at {station_name} ({time_str} vs prev {last_time_obj})")
                            
                    last_time_obj = current_time_obj
                    last_station_name = station_name
                except:
                    pass

                if is_new_trip:
                    split_count += 1
                    new_number = f"{train_number}_{split_count + 1}"
                    
                    # Create new train for the return/next trip
                    current_train = Train.objects.create(
                        number=new_number,
                        route=route,
                        days_operational=days
                    )
                    print(f"      ✨ Created split train {new_number}")
                    stops_created = 0 # Reset sequence for new train

                Stop.objects.create(
                    train=current_train,
                    station=station_objs[station_name],
                    departure_time=time_str,
                    sequence=stops_created + 1
                )
                stops_created += 1
            
            if skipped_stations:
                print(f"      ⚠️  Unknown stations: {', '.join(skipped_stations)}")
            
            if stops_created > 0:
                print(f"      ✅ Added {stops_created} stops to train {train_number}")
                imported_count += 1
            else:
                print(f"      ⚠️  No stops added for train {train_number}")
                skipped_count += 1
    
    print(f"\n✅ Import complete!")
    print(f"   Imported: {imported_count} trains")
    print(f"   Skipped: {skipped_count} trains")
    print(f"\n📊 Database stats:")
    print(f"   Total Trains: {Train.objects.count()}")
    print(f"   Total Stops: {Stop.objects.count()}")

if __name__ == "__main__":
    excel_path = "/home/mathxro/AntiGrav/SNTF_real.xlsx"
    
    print("=" * 60)
    print("  SNTF Train Data Import from Excel")
    print("=" * 60)
    
    import_excel_trains(excel_path)
    
    print("\n" + "=" * 60)
