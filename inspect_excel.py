import openpyxl
import pandas as pd

xls = pd.ExcelFile('SNTF_real.xlsx')

if 'Alger-Thenia' in xls.sheet_names:
    print("\n--- Inspecting Alger-Thenia Sheet ---")
    df = pd.read_excel(xls, 'Alger-Thenia', header=None)
    
    # Print first few rows to see structure
    print("First 10 rows:")
    print(df.head(10))
    
    # Check specific trains
    print("\nChecking Train 1025 (Column 2?):")
    print(df.iloc[:, 2].head(10))

# Original openpyxl code for Alger-Blida sheet
wb = openpyxl.load_workbook('SNTF_real.xlsx', data_only=True)
if 'Alger-Blida' in wb.sheetnames:
    ws = wb['Alger-Blida']
    print("\nFirst few stations from Alger-Blida:")
    for row in ws.iter_rows(min_row=7, max_row=12, values_only=True):
        print(row[:5])
